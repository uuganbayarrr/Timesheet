from datetime import datetime
import io, re, zipfile, json, subprocess, tempfile, pathlib
from collections import defaultdict
from typing import Optional, Tuple, List, Dict
import pdfplumber
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
from openpyxl.drawing.image import Image as XLImage
from pdf2image import convert_from_bytes
import os
import pandas as pd
import pytesseract
from PIL import Image
from pypdf import PdfWriter, PdfReader

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
logo_path = os.path.join(BASE_DIR, 'novelsoft_logo.png')

# ═══════════════════════════════════════════════════════════
# LOAD CONFIG FROM JSON
# ═══════════════════════════════════════════════════════════
def load_config(path="employees_config.json"):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def build_employee_lookup(config):
    return {str(e["sap_number"]): e for e in config.get("employees", [])}

def build_employee_lookup_by_name(config):
    """Case-insensitive full name → employee config row."""
    d = {}
    for e in config.get("employees", []):
        name = (e.get("employee_name") or "").strip().lower()
        if name:
            d[name] = e
    return d

# ═══════════════════════════════════════════════════════════
# STYLE HELPERS
# ═══════════════════════════════════════════════════════════
WHITE      = "FFFFFF"
BLACK      = "000000"
GRAY       = "D9D9D9"
LIGHT_BLUE = "DCE6F1"
MID_BLUE   = "4F81BD"
GREY       = "F2F2F2"
DATE_BOX_GREEN = "00B050"

def _s(style="thin", color=BLACK):   return Side(style=style, color=color)

def _b(left="thin", right="thin", top="thin", bottom="thin"):
    return Border(
        left=_s(left) if left else Side(style=None),
        right=_s(right) if right else Side(style=None),
        top=_s(top) if top else Side(style=None),
        bottom=_s(bottom) if bottom else Side(style=None),
    )
def _bm(style="thin"):  return Border(bottom=_s(style))
def _bdashed():
    return Border(left=_s("dashed"), right=_s("dashed"),
                  top=_s("dashed"), bottom=_s("dashed"))
def _dotted_border():
    return Border(left=_s("dotted"), right=_s("dotted"),
                  top=_s("dotted"), bottom=_s("dotted"))
def _green_box_border():
    g = DATE_BOX_GREEN
    return Border(
        left=_s("thin", g), right=_s("thin", g),
        top=_s("thin", g), bottom=_s("thin", g),
    )

def _fill(c): return PatternFill("solid", fgColor=c)

def _font(size=10, bold=False, color=BLACK, name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color)
def _al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sc(ws, r, c, value="", bold=False, size=10, color=BLACK, bg=None,
       h="left", v="center", wrap=False, bdr=None, nf=None):
    cell = ws.cell(r, c, value)
    cell.font = _font(size=size, bold=bold, color=color)
    cell.alignment = _al(h=h, v=v, wrap=wrap)
    if bg:  cell.fill   = _fill(bg)
    if bdr: cell.border = bdr
    if nf:  cell.number_format = nf
    return cell

def mc(ws, r1, c1, r2, c2, value="", bold=False, size=10, color=BLACK,
       bg=None, h="left", v="center", wrap=False, bdr=None, nf=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1, value)
    cell.font = _font(size=size, bold=bold, color=color)
    cell.alignment = _al(h=h, v=v, wrap=wrap)
    if bdr:
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(rr, cc).border = bdr
                if bg: ws.cell(rr, cc).fill = _fill(bg)
    elif bg:
        cell.fill = _fill(bg)
    if nf: cell.number_format = nf
    return cell

MONEY = '#,##0.00 "₮"'

# ═══════════════════════════════════════════════════════════
# PDF MERGE HELPER
# ═══════════════════════════════════════════════════════════
def merge_pdfs(pdf_bytes_list: list) -> bytes:
    """Хэд хэдэн PDF bytes-ийг дараалан нэгтгэж нэг PDF bytes буцаана."""
    writer = PdfWriter()
    for b in pdf_bytes_list:
        if not b:
            continue
        reader = PdfReader(io.BytesIO(b))
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()

# ═══════════════════════════════════════════════════════════
# XLSX → PDF bytes via LibreOffice
# ═══════════════════════════════════════════════════════════
def wb_to_pdf_bytes(wb):
    lo_bin = os.getenv("SOFFICE_BIN", "soffice")

    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = pathlib.Path(tmp) / "output.xlsx"
        wb.save(xlsx_path)
        subprocess.run(
            [lo_bin, "--headless", "--convert-to", "pdf", "--outdir", tmp, str(xlsx_path)],
            check=True,
            capture_output=True,
        )
        return xlsx_path.with_suffix(".pdf").read_bytes()

# ═══════════════════════════════════════════════════════════
# PDF PARSING
# ═══════════════════════════════════════════════════════════
def grab(pattern, text, default=""):
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(1).strip() if m else default

def parse_header(text):
    d = {}
    d["period_start"]   = grab(r"Period Start Date\s+([\d./-]+)", text)
    d["period_end"]     = grab(r"Period End Date\s+([\d./-]+)", text)
    d["sap_number"]     = grab(r"Number\s+(\d{5,})", text)
    d["employee_name"]  = grab(r"Employee Name/Surname\s+(.+?)\s+Total Number of Hours", text)
    d["position"]       = grab(r"(NOVELSOFT[^\n]+)", text)
    if not d["position"]:
        d["position"]   = grab(r"Position\s*/\s*Department\s+([A-Za-z][\w\s,]+?)\s+\*{2}", text)
    d["roster"]         = grab(r"Roster:\s+(.+?)\s+Number of Work Days", text)
    d["days_in_period"] = grab(r"Number of Days in Period\s+([\d.]+)", text)
    d["work_days"]      = grab(r"Number of Work Days\s+([\d.]+)", text)
    d["total_hours"]    = grab(r"Total Number of Hours\s+([\d.]+)", text)
    cc = re.findall(r"\b(490\d{5})\b", text)
    d["cost_code"]      = cc[0] if cc else ""
    return d

def parse_rows(text):
    row_re = re.compile(
        r"(\d{1,2}-\w{3})\s+(\S+(?:\s*/\s*\S+)?)\s+"
        r"([\d:]+|0:00|-)\s+([\d:]+|0:00|-)\s+"
        r"([\d.]+|-)\s+([\d.]+|-)\s*"
        r"([\d.]+)?\s*([\d.]+)?\s*(\d{6,})?\s*(.*)"
    )
    rows = []
    for line in text.splitlines():
        m = row_re.match(line.strip())
        if m:
            rows.append({
                "Date": m.group(1), "Day": m.group(2),
                "Time IN": m.group(3), "Time OUT": m.group(4),
                "Hours Worked": m.group(5), "Unpaid Break": m.group(6),
                "Overtime": m.group(7) or "", "Roster Break": m.group(8) or "",
                "Cost Code": m.group(9) or "", "Comment": m.group(10).strip(),
            })
    return rows

def parse_pdf(file_bytes):
    employees = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "MONTHLY EMPLOYEE TIMESHEET" in text:
                h = parse_header(text)
                r = parse_rows(text)
                if h.get("employee_name") or r:
                    employees.append({"header": h, "rows": r})

    if employees:
        return employees

    images = convert_from_bytes(file_bytes, dpi=300)
    for img in images:
        try:
            text = pytesseract.image_to_string(img, lang="mon+eng")
        except pytesseract.TesseractError:
            text = pytesseract.image_to_string(img, lang="eng")

        if "MONTHLY EMPLOYEE TIMESHEET" not in text:
            continue
        h = parse_header(text)
        r = parse_rows(text)
        if h.get("employee_name") or r:
            employees.append({"header": h, "rows": r})

    return employees

def enrich_header(header, emp_lookup):
    sap = str(header.get("sap_number", ""))
    cfg = emp_lookup.get(sap, {})

    if cfg.get("cost_code"):
        header["cost_code"] = cfg["cost_code"]
    elif not header.get("cost_code"):
        header["cost_code"] = "49071226"

    header["po_code"]        = cfg.get("po_code", "3106789606")
    header["ot_leader_name"] = cfg.get("ot_leader_name", "Munkhbayar Mishig")
    header["product_code"]   = cfg.get("product_code", "")

    if cfg.get("level"):
        header["level"] = cfg["level"]
    else:
        pos = (header.get("position") or "").lower()
        header["level"] = "senior" if "senior" in pos else "junior"

    # Convert total_hours to float
    header["total_hours"] = float(header.get("total_hours", 0) or 0)

    return header

# ═══════════════════════════════════════════════════════════
# GROUP EMPLOYEES BY PO CODE
# ═══════════════════════════════════════════════════════════
def group_by_po(emp_headers):
    groups = defaultdict(list)
    for h in emp_headers:
        po = h.get("po_code") or "UNKNOWN"
        groups[po].append(h)
    return dict(groups)

# ═══════════════════════════════════════════════════════════
# PRICING HELPERS
# ═══════════════════════════════════════════════════════════
def unit_price(header, pricing):
    level = header.get("level", "junior")
    return pricing.get("senior_unit_price", 32500) if level == "senior" else pricing.get("junior_unit_price", 25200)

def pos_label(header):
    level = header.get("level", "junior")
    return "Tableau Senior Expert" if level == "senior" else "Tableau Junior Expert"

ITEM_CODE_DEFAULT = {"senior": "7016003003200", "junior": "7016003020500"}

def k2_zarlaga_row_label(level: str) -> str:
    """Зарлагын баримтын барааны нэр — жишээ баримттай ижил."""
    return "K2 Developers - Senior" if (level or "").lower() == "senior" else "K2 Developers - Junior"


def item_code(header):
    if header.get("product_code"):
        return header["product_code"]
    level = header.get("level", "junior")
    return ITEM_CODE_DEFAULT.get(level, "7016003020500")

# ═══════════════════════════════════════════════════════════
# НЭХЭМЖЛЭХ  (Invoice) — returns PDF bytes
# ═══════════════════════════════════════════════════════════
def build_invoice(emp_list, pricing, company, doc_number="", doc_date="2/25/2026"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Нэхэмжлэх"
    ws.sheet_view.showGridLines = False

    widths = {1: 5, 2: 35, 3: 7, 4: 8, 5: 14, 6: 15, 7: 15}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.page_setup.orientation        = "portrait"
    ws.page_setup.paperSize           = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage           = True
    ws.page_setup.fitToWidth          = 1
    ws.page_setup.fitToHeight         = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left              = 0.5
    ws.page_margins.right             = 0.5
    ws.page_margins.top               = 0.5
    ws.page_margins.bottom            = 0.5

    R = 1

    if logo_path and os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 160
        img.height = 30
        img.anchor = "F2"
        ws.add_image(img)

        ws.merge_cells("F2:G2")
        ws["F2"] = ""

    # Continue with the rest of your layout
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 5, value="НовелСофт ХХК", bold=True, size=12)
    R += 1

    ws.row_dimensions[R].height = 30
    mc(ws, R, 1, R, 4,
       value="Новелсофт ХХК- Хаан банк цамхаг 22 давхар, Чингисийн өргөн чөлөө 6, Стадион оргил 1, Хан-Уул дүүрэг, Улаанбаатар 17010, Монгол улс",
       size=9, wrap=True)
    R += 1

    ws.row_dimensions[R].height = 16
    mc(ws, R, 1, R, 5, value="УТАС: (976)-72222828-3; WWW.NOVELSOFT.MN", size=9)
    R += 1
    ws.row_dimensions[R].height = 8
    R += 1

    ws.row_dimensions[R].height = 8
    R += 1

    ws.row_dimensions[R].height = 22
    mc(ws, R, 1, R, 3, value="", bdr=Border(bottom=_s("medium", MID_BLUE)))
    cell = mc(ws, R, 4, R, 5, value="НЭХЭМЖЛЭХ", bold=True, size=14, h="center")
    cell.font = Font(name="Times New Roman",bold=True, italic=True, size=14)
    mc(ws, R, 6, R, 7, value="", bdr=Border(bottom=_s("medium", MID_BLUE)))
    R += 1

    label_border = _bdashed()
    val_border   = _bdashed()

    po = emp_list[0].get("po_code", "3106789606") if emp_list else "3106789606"

    ws.row_dimensions[R].height = 22
    sc(ws, R, 1, "ХЭНД:", bold=True, size=9, bdr=label_border)
    mc(ws, R, 2, R, 4,
       value=company.get("invoice_recipient", "Оюу толгой ХХК"),
       bold=False, size=10, bdr=val_border)
    sc(ws, R, 5, "ДУГААР:", bold=True, size=9, bdr=label_border)
    mc(ws, R, 6, R, 7, value=doc_number, size=10, h="center", bdr=val_border)
    R += 1

    ws.row_dimensions[R].height = 46
    sc(ws, R, 1, "ХАЯГ:", bold=True, size=9, v="top", bdr=label_border)
    mc(ws, R, 2, R, 4,
       value=company.get("recipient_address", ""),
       size=9, wrap=True, v="top", bdr=val_border)
    sc(ws, R, 5, "ОГНОО:", bold=True, size=9, bdr=label_border)
    mc(ws, R, 6, R, 7, value=doc_date, size=10, h="center", bdr=val_border)
    R += 1

    ws.row_dimensions[R].height = 18
    sc(ws, R, 1, "УТАС:", bold=True, size=9, bdr=label_border)
    mc(ws, R, 2, R, 4, value=company.get("recipient_phone", ""), size=9, bdr=val_border)
    sc(ws, R, 5, "РО:", bold=True, size=10, bdr=label_border)
    mc(ws, R, 6, R, 7, value=po, bold=False, size=10, h="center", bdr=val_border)
    R += 1

    ws.row_dimensions[R].height = 8
    R += 1

    ws.row_dimensions[R].height = 26
    sc(ws, R, 1, "Д/д", bold=False, size=9, bg=GRAY, h="center", bdr=_b())
    sc(ws, R, 2, "Барааны нэр", bold=False, size=9, bg=GRAY, h="center", bdr=_b())
    mc(ws, R, 3, R, 4, value="Тоо", bold=False, size=9, bg=GRAY, h="center", bdr=_b())
    sc(ws, R, 5, "Нэгж үнэ", bold=False, size=9, bg=GRAY, h="center", bdr=_b())
    sc(ws, R, 6, "Нийт үнэ", bold=False, size=9, bg=GRAY, h="center", bdr=_b())
    sc(ws, R, 7, "Эцсийн үнэ", bold=False, size=9, bg=GRAY, h="center", bdr=_b())

    data_start = R + 1

    for idx, emp in enumerate(emp_list, 1):
        R += 1
        ws.row_dimensions[R].height = 42

        name = emp.get("employee_name", "")
        try:
            hours = float(emp.get("total_hours", 0))
        except Exception:
            hours = 0

        cc     = emp.get("cost_code", "49071226")
        per    = (emp.get("period_end") or "")[:7].replace(".", "-")
        leader = emp.get("ot_leader_name", "Munkhbayar Mishig")
        po_val = emp.get("po_code", po)
        u      = unit_price(emp, pricing)
        level = emp.get('level', 'junior')

        desc = (
            f"Tableau {level}/{per} сар {name} РО: {po_val}  OT\n"
            f"Asset management {leader} cost:{cc}"
        )

        sc(ws, R, 1, idx, size=9, h="center", v="top", bdr=_b())
        sc(ws, R, 2, desc, size=8.5, wrap=True, h="center", v="center", bdr=_b())
        mc(ws, R, 3, R, 4, value=hours, size=9, h="center", bdr=_b())
        sc(ws, R, 5, u, size=9, h="right", bdr=_b(), nf=MONEY)

        c6 = ws.cell(R, 6, f"=C{R}*E{R}")
        c6.font = _font(size=9); c6.alignment = _al(h="right")
        c6.border = _b(); c6.number_format = MONEY

        c7 = ws.cell(R, 7, f"=F{R}*1.1")
        c7.font = _font(size=9); c7.alignment = _al(h="right")
        c7.border = _b(); c7.number_format = MONEY

    data_end = R

    R += 1
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 6, value="Нийт төлбөр /НӨАТ ороогүй/ ₮",
    bold=False, size=9, h="right", bdr=_b())
    c = ws.cell(R, 7, f"=SUM(F{data_start}:F{data_end})")
    c.font = _font(size=9); c.alignment = _al(h="right")
    c.border = _b(); c.number_format = MONEY

    R += 1
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 6, value="НӨАТ 10% ₮",
    bold=False, size=9, h="right", bdr=_b())
    c = ws.cell(R, 7, f"=SUM(G{data_start}:G{data_end})-SUM(F{data_start}:F{data_end})")
    c.font = _font(size=9); c.alignment = _al(h="right")
    c.border = _b(); c.number_format = MONEY

    R += 1
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 6, value="Нийт төлбөр /НӨАТ орсон/ ₮", bold=True, size=9, h="right", bdr=_b())
    c = ws.cell(R, 7, f"=G{R-2}+G{R-1}")
    c.font = _font(size=9, bold=True); c.alignment = _al(h="right")
    c.border = _b(); c.number_format = MONEY

    R += 2
    ws.row_dimensions[R].height = 18
    mc(ws, R, 2, R, 4, value="БАТАЛГААЖУУЛСАН:", size=9, h="center")
    seller = company.get("seller_name", "Л. Анужин")

    R += 1
    ws.row_dimensions[R].height = 18
    sc(ws, R, 3, "/", size=9, h="center")
    mc(ws, R, 4, R, 5, value=seller, size=9, h="center", bdr=_bm())
    sc(ws, R, 6, "/", size=9, h="center")

    R += 3
    ws.row_dimensions[R].height = 16
    mc(ws, R, 1, R, 7,
       value="Гүйлгээний утга дээр компанийн нэр болон регистрийн дугаарыг заавал бичнэ үү.",
       size=9, h="center")

    R += 1
    ws.row_dimensions[R].height = 16
    mc(ws, R, 3, R, 5, value="Банкны мэдээлэл", bold=True, size=9, h="center")

    R += 1
    ws.row_dimensions[R].height = 22
    mc(ws, R, 1, R + 1, 2,
       value="Хүлээн авагчийн нэр:\nНовел Софт ХХК",
       bold=True, size=9, wrap=True, v="center")
    sc(ws, R, 3, company.get("bank_name", "ХХБанк"), size=9, h="center")
    mc(ws, R, 4, R, 5, value=company.get("bank_account", ""), size=9, h="center")
    sc(ws, R, 6, "Борлуулагч:", size=9, bold=False, h="right")
    sc(ws, R, 7, seller, size=9)

    R += 1
    ws.row_dimensions[R].height = 18
    sc(ws, R, 6, "Оффисын утас:", size=9, h="right")
    sc(ws, R, 7, company.get("seller_office_phone", "72222828"), size=9)

    R += 1
    ws.row_dimensions[R].height = 18
    sc(ws, R, 6, "Гар утас:", size=9, h="right")
    sc(ws, R, 7, company.get("seller_mobile", "95001168"), size=9)

    R += 1
    ws.row_dimensions[R].height = 18
    sc(ws, R, 6, "И-мэйл:", size=9, h="right")
    sc(ws, R, 7, company.get("seller_email", "anujin.l@novelsoft.mn"), size=9, color=MID_BLUE)

    ws.print_area = f"A1:G{R}"

    return wb_to_pdf_bytes(wb)

# ═══════════════════════════════════════════════════════════
# ЗАРЛАГЫН БАРИМТ — returns PDF bytes
# ═══════════════════════════════════════════════════════════
def build_zarlagiin(emp_list, pricing, company, doc_number="", doc_date=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Зарлагын баримт"
    ws.sheet_view.showGridLines = False

    widths = {1: 4.5, 2: 18, 3: 28, 4: 10, 5: 22, 6: 22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right  = 0.5
    ws.page_margins.top  = ws.page_margins.bottom = 0.5

    R = 1

    if logo_path and os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 160
        img.height = 30
        img.anchor = "F1"
        ws.add_image(img)

    ws.row_dimensions[R].height = 16
    mc(ws, R, 1, R, 3, value="НовелСофт ХХК", bold=True, size=12)
    R += 1
    ws.row_dimensions[R].height = 30
    mc(ws, R, 1, R, 3,
       value="Новелсофт ХХК- Хаан банк цамхаг 22 давхар, Чингисийн өргөн чөлөө 6, Стадион оргил 1, Хан-Уул дүүрэг, Улаанбаатар 17010, Монгол улс",
       size=9, wrap=True)
    R += 1

    ws.row_dimensions[R].height = 15
    mc(ws, R, 1, R, 3, value="УТАС: (976)-72222828-3; WWW.NOVELSOFT.MN", size=9)
    R += 1

    ws.row_dimensions[R].height = 6
    R += 1

    ws.row_dimensions[R].height = 22
    mc(ws, R, 1, R, 3, value="", bdr=Border(bottom=_s("medium", MID_BLUE)))
    cell = mc(ws, R, 4, R, 5, value="ЗАРЛАГЫН БАРИМТ", bold=True, size=14, h="center")
    cell.font = Font(name="Times New Roman", bold=True, italic=True, size=14)
    mc(ws, R, 6, R, 6, value="", bdr=Border(bottom=_s("medium", MID_BLUE)))
    R += 1

    info_border = _dotted_border()
    gdate = _green_box_border()

    ws.row_dimensions[R].height = 44
    mc(ws, R, 1, R, 2, value="Хүлээн авагчийн нэр:", size=9, bdr=info_border)
    sc(ws, R, 3, company.get("invoice_recipient", "Оюу толгой ХХК"),
       bold=True, size=10, h="center", bdr=info_border)
    sc(ws, R, 4, "Дугаар:", size=9, bdr=info_border)
    mc(ws, R, 5, R, 6, value=doc_number, size=10, h="center", bdr=info_border)
    R += 1

    ws.row_dimensions[R].height = 44
    mc(ws, R, 1, R, 2, value="Хүргэх хаяг:", size=9, v="top", bdr=info_border)
    sc(ws, R, 3, company.get("recipient_address", ""),
       size=9, wrap=True, v="top", bdr=info_border)
    sc(ws, R, 4, "Огноо:", size=9, bdr=info_border)
    mc(ws, R, 5, R, 6, value=doc_date, size=10, h="center", bdr=info_border)
    for cc in range(5, 7):
        ws.cell(R, cc).border = gdate
    R += 1

    ws.row_dimensions[R].height = 14
    R += 1

    ws.row_dimensions[R].height = 28
    header_bg = "D9D9D9"
    sc(ws, R, 1, "Д/д",        bold=False, size=9, bg=header_bg, h="center", bdr=_b())
    sc(ws, R, 2, "Барааны код", bold=False, size=9, bg=header_bg, h="center", bdr=_b())
    sc(ws, R, 3, "Барааны нэр", bold=False, size=9, bg=header_bg, h="center", bdr=_b())
    sc(ws, R, 4, "Тоо",         bold=False, size=9, bg=header_bg, h="center", bdr=_b())
    sc(ws, R, 5, "Нэгж үнэ",   bold=False, size=9, bg=header_bg, h="center", bdr=_b())
    sc(ws, R, 6, "Нийт үнэ",   bold=False, size=9, bg=header_bg, h="center", bdr=_b())

    DS = R + 1

    for idx, emp in enumerate(emp_list, 1):
        R += 1
        ws.row_dimensions[R].height = 42 if len(str(pos_label(emp))) > 20 else 28

        qty = emp.get("total_hours", 0)
        try: qty = float(qty)
        except: qty = 0

        u     = unit_price(emp, pricing)
        label = pos_label(emp)
        code  = item_code(emp)

        sc(ws, R, 1, idx,   size=9, h="center", v="center", bdr=_b())
        sc(ws, R, 2, code,  size=9, h="center", v="center", bdr=_b())
        sc(ws, R, 3, label, size=9, h="center", v="center", wrap=True, bdr=_b())
        sc(ws, R, 4, qty,   size=9, h="center", v="center", bdr=_b())
        sc(ws, R, 5, u,     size=9, h="right",  v="center", bdr=_b(), nf=MONEY)

        c = ws.cell(R, 6, f"=D{R}*E{R}")
        c.font = _font(size=9); c.alignment = _al(h="right", v="center")
        c.border = _b(); c.number_format = MONEY

    DE = R

    R += 1
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 5, value="Нийт төлбөр /НӨАТ ороогүй/ ₮",
       bold=False, size=9, h="right", bdr=_b())
    c = ws.cell(R, 6, f"=SUM(F{DS}:F{DE})")
    c.number_format = MONEY; c.border = _b()
    c.alignment = _al(h="right"); c.font = _font(size=9)

    R += 1
    nv = R
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 5, value="НӨАТ 10% ₮", bold=False, size=9, h="right", bdr=_b())
    c = ws.cell(R, 6, f"=F{R-1}*0.1")
    c.number_format = MONEY; c.border = _b()
    c.alignment = _al(h="right"); c.font = _font(size=9)

    R += 1
    ws.row_dimensions[R].height = 20
    mc(ws, R, 1, R, 5, value="Нийт төлбөр /НӨАТ орсон/ ₮",
       bold=True, size=10, h="right", bdr=_b())
    c = ws.cell(R, 6, f"=F{nv-1}+F{nv}")
    c.number_format = MONEY; c.border = _b()
    c.alignment = _al(h="right"); c.font = _font(size=10, bold=True)

    R += 1
    ws.row_dimensions[R].height = 6
    for cc in range(1, 7):
        ws.cell(R, cc).border = Border(top=_s("medium", BLACK))

    seller = "/" + company.get("seller_name", "Л. Анужин") + "/"

    R += 1
    ws.row_dimensions[R].height = 15

    R += 1
    ws.row_dimensions[R].height = 20
    sc(ws, R, 1, "Борлуулагч:", size=9)
    sc(ws, R, 2, "", bdr=_bm())
    sc(ws, R, 3, seller, size=9, h="center")
    sc(ws, R, 4, "Барааг хүргэсэн түгээгч:", size=9)
    sc(ws, R, 5, "", bdr=_bm())
    sc(ws, R, 6, "", bdr=_bm())

    R += 1
    ws.row_dimensions[R].height = 40
    sc(ws, R, 1, value="Барааг олгосон нярав:", size=9)
    sc(ws, R, 2, value="", bdr=_bm())
    sc(ws, R, 3, value="")
    mc(ws, R, 4, R, 6,
       value="Бараа/ хайрцаг сав баглаа боодол\nбүрэн шалгаж/ хүлээн авсан:",
       size=9, wrap=True, v="center", h="left")
    R += 1
    sc(ws, R, 5, value="", bdr=_bm())
    sc(ws, R, 6, value="", bdr=_bm())

    return wb_to_pdf_bytes(wb)

# ═══════════════════════════════════════════════════════════
# DAILY TIMESHEET EXCEL → PDF  (per employee)
# ═══════════════════════════════════════════════════════════
COL_KEYS = ["Date","Day","Time IN","Time OUT","Hours Worked",
            "Unpaid Break","Overtime","Roster Break","Cost Code","Comment"]

def wb_bytes(wb):
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def safe(name):
    return re.sub(r"[^\w\s-]", "", name).replace(" ", "_")

def build_timesheet_pdf(header: dict, rows: list) -> bytes:
    """Parse хийсэн timesheet мөрийг хүснэгт PDF болгоно (Нэхэмжлэх/Зарлагаас тусдаа)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.sheet_view.showGridLines = True
    r = 1
    for label, key in [
        ("Employee", "employee_name"),
        ("SAP", "sap_number"),
        ("Period Start", "period_start"),
        ("Period End", "period_end"),
        ("PO", "po_code"),
        ("Total Hours", "total_hours"),
    ]:
        sc(ws, r, 1, label, bold=True, size=9, bdr=_b())
        sc(ws, r, 2, str(header.get(key, "") or ""), size=9, bdr=_b())
        r += 1
    r += 1
    for c, k in enumerate(COL_KEYS, 1):
        sc(ws, r, c, k, bold=True, size=9, bg=GRAY, h="center", bdr=_b())
    r += 1
    for row in rows:
        for c, k in enumerate(COL_KEYS, 1):
            sc(ws, r, c, row.get(k, ""), size=9, bdr=_b())
        r += 1
    return wb_to_pdf_bytes(wb)

def enrich_project_list_employee(emp: dict, name_lookup: dict, pricing: dict) -> dict:
    """employees_config.json-оос level, po_code зэргийг нэмэх (Хэрэглээтэй бол)."""
    emp = dict(emp)
    key = emp.get("employee_name", "").strip().lower()
    cfg = name_lookup.get(key)
    if cfg:
        if cfg.get("level"):
            emp["level"] = cfg["level"]
        if cfg.get("cost_code"):
            emp["cost_code"] = cfg["cost_code"]
        if cfg.get("po_code") and not emp.get("po_code"):
            emp["po_code"] = cfg["po_code"]
    emp["employee_name"] = strip_time_from_text(emp.get("employee_name", ""))
    lv = (emp.get("level") or "junior").lower()
    up = float(
        pricing.get("senior_unit_price", 32500)
        if lv == "senior"
        else pricing.get("junior_unit_price", 25200)
    )
    emp["unit_price"] = up
    return emp

def order_special_employees(special_raw: List[Dict]) -> List[Dict]:
    """Shajinbat → Sergelen дараалал (хоёуланг нь PDF-д оруулах)."""
    order_keys = ("shajinbat", "sergelen")
    by_key = {e["employee_name"].strip().lower(): e for e in special_raw}
    return [by_key[k] for k in order_keys if k in by_key]


def excel_po_k2_output_files(
    po_code: str,
    po_data: dict,
    company: dict,
    pricing: dict,
    name_lookup: dict,
    meta: dict,
) -> List[Tuple[str, bytes]]:
    """
    Нэгтгэсэн PDF хуудасны дараалал (PO тус бүр):
      Хуудас 1: Бусад ажилтны нэхэмжлэх (Senior, дараа нь Developer)
      Хуудас 2: Бусад ажилтны зарлагын баримт
      Хуудас 3: Shajinbat нэхэмжлэх (байвал)b
      Хуудас 4: Shajinbat зарлагын баримт (байвал)
      Хуудас 5: Sergelen нэхэмжлэх (байвал)
      Хуудас 6: Sergelen зарлагын баримт (байвал)
    """
    y, m = period_ym_from_doc_date(meta.get("doc_date") or "")
    sub = k2_subheader_text(y, m, po_code)
    enriched = [
        enrich_project_list_employee(dict(e), name_lookup, pricing)
        for e in po_data["employees"]
    ]
    main_raw, special_raw = split_excel_employees_for_k2(enriched)
    ordered_special = order_special_employees(special_raw)

    out: List[Tuple[str, bytes]] = []

    if main_raw:
        agg = aggregate_main_excel_by_senior_junior(main_raw, pricing)
        if agg:
            fill_k2_invoice_descriptions(agg, y, m, po_code)
            inv = build_project_invoice(
                po_code, {"employees": agg}, company,
                meta["inv_number"], meta["doc_date"],
                k2_subheader=sub, skip_grouping=True, employee_name=[e["employee_name"] for e in main_raw if e.get("employee_name")][0]
            )
            zar = build_project_zarlagiin(
                po_code, {"employees": agg}, company,
                meta["zar_number"], meta["doc_date"], employee_name=[e["employee_name"] for e in main_raw if e.get("employee_name")][0]
            )
            stem = f"{safe(po_code)}_K2_bussad"
            out.append((stem, merge_pdfs([inv, zar])))

    for sp in ordered_special:
        row = dict(sp)
        fill_k2_invoice_descriptions([row], y, m, po_code)
        inv = build_project_invoice(
            po_code, {"employees": [row]}, company,
            meta["inv_number"], meta["doc_date"],
            k2_subheader=sub, skip_grouping=True,employee_name=[e["employee_name"] for e in main_raw if e.get("employee_name")][0]
        )
        # Pass `row` (not sp) so zarlaga also sees the enriched employee name
        zar = build_project_zarlagiin(
            po_code, {"employees": [dict(row)]}, company,
            meta["zar_number"], meta["doc_date"], employee_name= [e["employee_name"] for e in main_raw if e.get("employee_name")][0]
        )
        stem = f"{safe(po_code)}_{safe(sp.get('employee_name', 'special'))}"
        out.append((stem, merge_pdfs([inv, zar])))

    return out

# ═══════════════════════════════════════════════════════════
# PARSE PROJECT LIST EXCEL
# ═══════════════════════════════════════════════════════════
def parse_project_list_excel(file_bytes: bytes) -> dict:
    """
    Reads ConT/VWL Project List Excel.
    Columns: ProjectName | Department | CostCenter | FullName | Service (PO) | WorkedHour | UnitPrice
    Returns dict keyed by PO code.
    """
    df = pd.read_excel(io.BytesIO(file_bytes), header=0)
    df.columns = ["ProjectName", "Department", "CostCenter",
                  "FullName", "Service", "WorkedHour", "UnitPrice"]

    po_pattern = re.compile(r"^31\d{8}$")

    def is_valid_po(x):
        try:
            return bool(po_pattern.match(str(x).strip()))
        except Exception:
            return False

    detail = df[
        df["ProjectName"].notna() &
        df["Service"].apply(is_valid_po)
    ].copy()

    detail["Service"]    = detail["Service"].astype(str).str.strip()
    detail["WorkedHour"] = pd.to_numeric(detail["WorkedHour"], errors="coerce").fillna(0)
    detail["UnitPrice"]  = pd.to_numeric(detail["UnitPrice"],  errors="coerce").fillna(0)
    detail["CostCenter"] = detail["CostCenter"].astype(str).str.strip().str.rstrip(".0")

    result = {}
    for po, po_df in detail.groupby("Service"):
        employees = []
        for (emp_name, unit_p), emp_df in po_df.groupby(["FullName", "UnitPrice"]):
            total_hours = float(emp_df["WorkedHour"].sum())
            cost_code   = str(emp_df["CostCenter"].iloc[0])
            level       = "senior" if unit_p >= 32500 else "junior"
            projects    = [
                {
                    "name":  str(row["ProjectName"]),
                    "dept":  str(row["Department"]),
                    "hours": float(row["WorkedHour"]),
                }
                for _, row in emp_df.iterrows()
            ]
            employees.append({
                "employee_name": strip_time_from_text(str(emp_name)),
                "total_hours":   total_hours,
                "unit_price":    float(unit_p),
                "cost_code":     cost_code,
                "level":         level,
                "projects":      projects,
                "po_code":       po,
            })
        result[po] = {"employees": employees}

    return result


# ═══════════════════════════════════════════════════════════
# BUILD PROJECT НЭХЭМЖЛЭХ PDF  (per PO from Excel)
# ═══════════════════════════════════════════════════════════
def build_project_invoice(
    po_code,
    po_data,
    company,
    doc_number="",
    doc_date="",
    k2_subheader: Optional[str] = None,
    skip_grouping: bool = False,
    employee_name: Optional[str] = None,
):
    raw_employees = po_data["employees"]
    employees = list(raw_employees) if skip_grouping else group_employees_for_invoice(raw_employees)

    wb = Workbook()
    ws = wb.active
    ws.title = "Нэхэмжлэх"
    ws.sheet_view.showGridLines = False

    # -----------------------------
    # COLUMN WIDTHS
    # -----------------------------
    # ### гарч байсан гол шалтгаан нь 5,6,7-р баганууд хэт нарийн байсан
    widths = {
        1: 5,    # A  Д/д
        2: 40,   # B  Барааны нэр / тайлбар
        3: 8,    # C  Тоо (merge start)
        4: 8,    # D  Тоо (merge end)
        5: 16,   # E  Нэгж үнэ
        6: 16,   # F  Нийт үнэ
        7: 16,   # G  Эцсийн үнэ
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # -----------------------------
    # PAGE SETUP
    # -----------------------------
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.print_options.gridLines = False

    R = 1

    # -----------------------------
    # LOGO
    # -----------------------------
    if logo_path and os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 160
        img.height = 30
        img.anchor = "F1"
        ws.add_image(img)

    # -----------------------------
    # HEADER
    # -----------------------------
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 4, value="Новел Софт", bold=True, size=12)
    R += 1

    ws.row_dimensions[R].height = 30
    mc(
        ws, R, 1, R, 4,
        value="НовелСофт ХХК, 22-р давхар, Хаан банк цамхаг; Чингисийн өргөн чөлөө-6, Стадион Оргил-1, Хан-Уул дүүрэг Улаанбаатар-17010,",
        size=9, wrap=True
    )
    R += 1

    ws.row_dimensions[R].height = 16
    mc(ws, R, 1, R, 5, value="УТАС: (976)-72222828-3; WWW.NOVELSOFT.MN", size=9)
    R += 1

    ws.row_dimensions[R].height = 8
    R += 1

    # -----------------------------
    # TITLE
    # -----------------------------
    ws.row_dimensions[R].height = 22
    mc(ws, R, 1, R, 3, value="", bdr=Border(bottom=_s("medium", MID_BLUE)))
    cell = mc(ws, R, 4, R, 5, value="НЭХЭМЖЛЭХ", bold=True, size=14, h="center")
    cell.font = Font(name="Times New Roman", bold=True, italic=True, size=14)
    mc(ws, R, 6, R, 7, value="", bdr=Border(bottom=_s("medium", MID_BLUE)))
    R += 1

    lb = _bdashed()
    vb = _bdashed()

    # -----------------------------
    # RECIPIENT BLOCK
    # -----------------------------
    ws.row_dimensions[R].height = 22
    sc(ws, R, 1, "ХЭНД:", bold=True, size=9, bdr=lb)
    mc(
        ws, R, 2, R, 4,
        value=company.get("invoice_recipient", "Оюу толгой ХХК"),
        bold=False, size=10, bdr=vb
    )
    sc(ws, R, 5, "ДУГААР:", bold=True, size=9, bdr=lb)
    mc(ws, R, 6, R, 7, value=doc_number, size=10, h="center", bdr=vb)
    R += 1

    ws.row_dimensions[R].height = 46
    sc(ws, R, 1, "ХАЯГ:", bold=True, size=9, v="top", bdr=lb)
    mc(
        ws, R, 2, R, 4,
        value=company.get("recipient_address", ""),
        size=9, wrap=True, v="top", bdr=vb
    )
    sc(ws, R, 5, "ОГНОО:", bold=True, size=9, bdr=lb)
    mc(ws, R, 6, R, 7, value=doc_date, size=10, h="center", bdr=vb)
    R += 1

    ws.row_dimensions[R].height = 18
    sc(ws, R, 1, "Утас:", bold=True, size=9, bdr=lb)
    mc(ws, R, 2, R, 4, value=company.get("recipient_phone", ""), size=9, bdr=vb)
    sc(ws, R, 5, "PO:", bold=True, size=10, bdr=lb)
    mc(ws, R, 6, R, 7, value=po_code, bold=True, size=10, h="center", bdr=vb)
    R += 1

    ws.row_dimensions[R].height = 8
    R += 1

    # -----------------------------
    # TABLE HEADER
    # -----------------------------
    ws.row_dimensions[R].height = 42
    sc(ws, R, 1, "Д/д", bold=False, size=9, bg=GRAY, h="center", v="center", bdr=_b())
    mc(
        ws, R, 2, R, 2,
        value=k2_subheader,
        bold=False, size=9, bg=GRAY,
        h="center", v="center", wrap=True, bdr=_b()
    )
    mc(ws, R, 3, R, 4, value="Тоо", bold=False, size=9, bg=GRAY, h="center", v="center", bdr=_b())
    sc(ws, R, 5, "Нэгж үнэ", bold=False, size=9, bg=GRAY, h="center", v="center", bdr=_b())
    sc(ws, R, 6, "Нийт үнэ", bold=False, size=9, bg=GRAY, h="center", v="center", bdr=_b())
    sc(ws, R, 7, "Эцсийн үнэ", bold=False, size=9, bg=GRAY, h="center", v="center", bdr=_b())

    data_start = R + 1

    # -----------------------------
    # TABLE ROWS
    # -----------------------------
    for idx, emp in enumerate(employees, 1):
        R += 1
        ws.row_dimensions[R].height = 56

        name = emp["employee_name"]
        hours = emp["total_hours"]
        u = emp["unit_price"]
        cc = emp["cost_code"]

        year, month = period_ym_from_doc_date(meta.get("doc_date") or "")
        if employee_name == "Sergelen Saranmandakh":
            desc = f" Software үйлчилгээний {year} {month}-р сарын төлбөр"
        elif employee_name == "Shajinbat Tsogbadrakh":
            desc = f"Sharepoint үйлчилгээний {year} {month}-р сарын төлбөр"
        elif emp.get("invoice_desc"):
            desc = emp["invoice_desc"]
        else:
            desc = f"K2 expert/{name}  РО: {po_code}\nCost code: {cc}"

        sc(ws, R, 1, idx, size=9, h="center", v="top", bdr=_b())
        sc(ws, R, 2, desc, size=8.5, wrap=True, h="center", v="center", bdr=_b())
        mc(ws, R, 3, R, 4, value=hours, size=9, h="center", v="center", bdr=_b())
        sc(ws, R, 5, u, size=9, h="right", v="center", bdr=_b(), nf=MONEY)

        c6 = ws.cell(R, 6, f"=C{R}*E{R}")
        c6.font = _font(size=9)
        c6.alignment = _al(h="right", v="center")
        c6.border = _b()
        c6.number_format = MONEY

        c7 = ws.cell(R, 7, f"=F{R}*1.1")
        c7.font = _font(size=9)
        c7.alignment = _al(h="right", v="center")
        c7.border = _b()
        c7.number_format = MONEY

    data_end = R

    # -----------------------------
    # TOTALS
    # -----------------------------
    R += 1
    ws.row_dimensions[R].height = 18
    mc(
        ws, R, 1, R, 6,
        value="Нийт төлбөр /НӨАТ ороогүй/ ₮",
        bold=False, size=9, h="right", bdr=_b()
    )
    c = ws.cell(R, 7, f"=SUM(F{data_start}:F{data_end})")
    c.font = _font(size=9)
    c.alignment = _al(h="right")
    c.border = _b()
    c.number_format = MONEY

    R += 1
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 6, value="НӨАТ 10% ₮", bold=False, size=9, h="right", bdr=_b())
    c = ws.cell(R, 7, f"=SUM(G{data_start}:G{data_end})-SUM(F{data_start}:F{data_end})")
    c.font = _font(size=9)
    c.alignment = _al(h="right")
    c.border = _b()
    c.number_format = MONEY

    R += 1
    ws.row_dimensions[R].height = 18
    mc(
        ws, R, 1, R, 6,
        value="Нийт төлбөр /НӨАТ орсон/ ₮",
        bold=True, size=9, h="right", bdr=_b()
    )
    c = ws.cell(R, 7, f"=G{R-2}+G{R-1}")
    c.font = _font(size=9, bold=True)
    c.alignment = _al(h="right")
    c.border = _b()
    c.number_format = MONEY

    seller = company.get("seller_name", "Л. Анужин")

    # -----------------------------
    # SIGNATURE
    # -----------------------------
    R += 2
    ws.row_dimensions[R].height = 18
    mc(ws, R, 2, R, 4, value="БАТАЛГААЖУУЛСАН:", size=9, h="center")

    R += 1
    ws.row_dimensions[R].height = 18
    sc(ws, R, 3, "/", size=9, h="center")
    mc(ws, R, 4, R, 5, value=seller, size=9, h="center", bdr=_bm())
    sc(ws, R, 6, "/", size=9, h="center")

    # -----------------------------
    # BANK INFO
    # -----------------------------
    R += 3
    mc(
        ws, R, 1, R, 7,
        value="Гүйлгээний утга дээр компанийн нэр болон регистрийн дугаарыг заавал бичнэ үү.",
        size=9, h="center"
    )

    R += 1
    mc(ws, R, 3, R, 5, value="Банкны мэдээлэл", bold=True, size=9, h="center")

    R += 1
    mc(
        ws, R, 1, R + 1, 2,
        value="Хүлээн авагчийн нэр:\nНовел Софт ХХК",
        bold=True, size=9, wrap=True, v="center"
    )
    sc(ws, R, 3, company.get("bank_name", "ХХБанк"), size=9, h="center")
    mc(ws, R, 4, R, 5, value=company.get("bank_account", ""), size=9, h="center")
    sc(ws, R, 6, "Борлуулагч:", size=9, h="right")
    sc(ws, R, 7, seller, size=9)

    R += 1
    sc(ws, R, 6, "Оффисын утас:", size=9, h="right")
    sc(ws, R, 7, company.get("seller_office_phone", "72222828"), size=9)

    R += 1
    sc(ws, R, 6, "Гар утас:", size=9, h="right")
    sc(ws, R, 7, company.get("seller_mobile", "95001168"), size=9)

    R += 1
    sc(ws, R, 6, "И-мэйл:", size=9, h="right")
    sc(ws, R, 7, company.get("seller_email", "anujin.l@novelsoft.mn"), size=9, color=MID_BLUE)

    # -----------------------------
    # PRINT AREA
    # -----------------------------
    ws.print_area = f"A1:G{R}"

    return wb_to_pdf_bytes(wb)

# ═══════════════════════════════════════════════════════════
# BUILD PROJECT ЗАРЛАГЫН БАРИМТ PDF
# ═══════════════════════════════════════════════════════════
def item_code(emp):
    return emp.get("item_code") or "7016003002700"

def strip_time_from_text(text):
    return text

def k2_zarlaga_row_label(level):
    level = (level or "").lower()
    if level == "senior":
        return "Software Developer Senior"
    if level == "junior":
        return "Software Developer Junior"
    return "Software Developer"

def build_project_zarlagiin(po_code, po_data, company, doc_number="", doc_date="", employee_name: Optional[str] = None):
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    employees = sorted(
        po_data["employees"],
        key=lambda e: 0 if (e.get("level") or "junior").lower() == "senior" else 1,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Зарлагын баримт"
    ws.sheet_view.showGridLines = False

    # =========================
    # BAGANA / ZURAG SHIG 7 COL
    # A  : Д/д
    # B  : Барааны код
    # C  : Барааны нэр
    # D  : Тоо
    # E  : Нэгж үнэ
    # F  : ₮
    # G  : Нийт үнэ
    # =========================
    widths = {
        1: 4.5,   # A
        2: 18,    # B
        3: 31,    # C
        4: 10,    # D
        5: 18,    # E
        6: 4.0,   # F -> ₮
        7: 18,    # G
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right  = 0.5
    ws.page_margins.top  = ws.page_margins.bottom = 0.5

    # -------- helpers --------
    BLACK = "000000"
    MID_BLUE = "4F81BD"
    LIGHT_GRAY = "D9D9D9"

    def _s(style="thin", color=BLACK):
        return Side(style=style, color=color)

    def _al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _font(size=10, bold=False, italic=False, name="Times New Roman"):
        return Font(name=name, size=size, bold=bold, italic=italic)

    def _b():
        return Border(left=_s("thin"), right=_s("thin"), top=_s("thin"), bottom=_s("thin"))

    def _bm():
        return Border(bottom=_s("thin"))

    def _dotted_border():
        return Border(
            left=_s("dotted"),
            right=_s("dotted"),
            top=_s("dotted"),
            bottom=_s("dotted"),
        )

    def _green_box_border():
        return Border(
            left=_s("dotted"),
            right=_s("dotted"),
            top=_s("dotted"),
            bottom=_s("dotted"),
        )

    def sc(ws_, r, c, value="", bold=False, size=10, italic=False,
           h="left", v="center", wrap=False, bg=None, bdr=None, nf=None):
        cell = ws_.cell(r, c, value)
        cell.font = _font(size=size, bold=bold, italic=italic)
        cell.alignment = _al(h=h, v=v, wrap=wrap)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if bdr:
            cell.border = bdr
        if nf:
            cell.number_format = nf
        return cell

    def mc(ws_, r1, c1, r2, c2, value="", bold=False, size=10, italic=False,
           h="center", v="center", wrap=False, bg=None, bdr=None, nf=None):
        ws_.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws_.cell(r1, c1, value)
        cell.font = _font(size=size, bold=bold, italic=italic)
        cell.alignment = _al(h=h, v=v, wrap=wrap)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if bdr:
            # merge range бүх cell-д border өгөх
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    ws_.cell(rr, cc).border = bdr
        if nf:
            cell.number_format = nf
        return cell

    MONEY = '#,##0.00₮'

    R = 1

    # ========= Header =========
    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 4, value="Новел Софт ХХК", bold=True, size=12, h="left")
    R += 1

    ws.row_dimensions[R].height = 22
    mc(
        ws, R, 1, R, 4,
        value="Новелсофт ХХК, 22-р давхар, Хаан банк цамхаг,\n"
              "Чингисийн өргөн чөлөө-6, Стадион Оргил-1,",
        size=9, h="left", wrap=True
    )
    R += 1

    ws.row_dimensions[R].height = 18
    mc(ws, R, 1, R, 4, value="УТАС: (976)-72222828-3; WWW.NOVELSOFT.MN", size=9, h="left")
    R += 1

    if logo_path and os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 160
        img.height = 30
        img.anchor = "F1"
        ws.add_image(img)

    # зураг дээрх хөх зураас
    ws.row_dimensions[R].height = 8
    for c in range(1, 8):
        ws.cell(R, c).border = Border(bottom=_s("medium", MID_BLUE))
    R += 1

    # Title row
    ws.row_dimensions[R].height = 26
    mc(ws, R, 1, R, 3, value="")
    title = mc(ws, R, 4, R, 5, value="ЗАРЛАГЫН БАРИМТ", bold=True, italic=True, size=15, h="center")
    title.font = Font(name="Times New Roman", size=15, bold=True, italic=True)
    mc(ws, R, 6, R, 7, value="")
    R += 1

    ib = _dotted_border()
    gdate = _green_box_border()

    # Recipient / Number
    ws.row_dimensions[R].height = 34
    sc(ws, R, 1, "Нэр:", size=9, bdr=ib)
    mc(
        ws, R, 2, R, 4,
        value=company.get("invoice_recipient", "Оюу толгой ХХК"),
        bold=True, size=10, h="center", bdr=ib
    )
    sc(ws, R, 5, "ДУГААР:", size=9, bdr=ib)
    mc(ws, R, 6, R, 7, value=doc_number, size=10, h="center", bdr=ib)
    R += 1

    # Address / Date
    ws.row_dimensions[R].height = 46
    sc(ws, R, 1, "Хаяг:", size=9, v="top", bdr=ib)
    mc(
        ws, R, 2, R, 4,
        value=company.get(
            "recipient_address",
            "Монгол улс, Улаанбаатар - 14240,\n"
            "Сүхбаатар дүүрэг, Чингисийн өргөн\n"
            "чөлөө - 15, “Моннис” цамхаг"
        ),
        size=9, wrap=True, h="left", v="top", bdr=ib
    )
    sc(ws, R, 5, "ОГНОО:", size=9, bdr=ib)
    mc(ws, R, 6, R, 7, value=doc_date, size=10, h="center", bdr=ib)
    for cc in range(6, 8):
        ws.cell(R, cc).border = gdate
    R += 1

    ws.row_dimensions[R].height = 16
    R += 1

    # ========= Table header =========
    ws.row_dimensions[R].height = 30
    sc(ws, R, 1, "Д/д", size=9, h="center", bg=LIGHT_GRAY, bdr=_b())
    sc(ws, R, 2, "Барааны код", size=9, h="center", bg=LIGHT_GRAY, bdr=_b())
    sc(ws, R, 3, "Барааны нэр", size=9, h="center", bg=LIGHT_GRAY, bdr=_b())
    sc(ws, R, 4, "Тоо", size=9, h="center", bg=LIGHT_GRAY, bdr=_b())
    sc(ws, R, 5, "Нэгж үнэ", size=9, h="center", bg=LIGHT_GRAY, bdr=_b())
    mc(ws, R, 6, R, 7, value="Нийт үнэ", size=9, h="center", bg=LIGHT_GRAY, bdr=_b())

    # ========= Detail rows =========
    # Зураг дээрх шиг 1 мөртэй харагдах байдлаар
    DS = R + 1

    for idx, emp in enumerate(employees, 1):
        R += 1
        ws.row_dimensions[R].height = 32

        qty = emp["total_hours"]
        unit_price = emp["unit_price"]
        # level = (emp.get("level") or "").lower()
        # emp_name = (emp.get("employee_name") or "").strip()

        year, month = period_ym_from_doc_date(meta.get("doc_date") or "")
        if employee_name == "Sergelen Saranmandakh":
            item_name = f" Software үйлчилгээний {year} {month}-р сарын төлбөр"
        elif employee_name == "Shajinbat Tsogbadrakh":
            item_name = f"Sharepoint үйлчилгээний {year} {month}-р сарын төлбөр"
        elif emp.get("invoice_desc"):
            item_name = emp["invoice_desc"]

        code = item_code(emp) if "item_code" in globals() else po_code

        sc(ws, R, 1, idx, size=9, h="center", bdr=_b())
        sc(ws, R, 2, code, size=9, h="center", bdr=_b())
        sc(ws, R, 3, item_name, size=9, h="center", wrap=True, bdr=_b())
        sc(ws, R, 4, qty, size=9, h="center", bdr=_b())
        sc(ws, R, 5, unit_price, size=9, h="right", bdr=_b(), nf=MONEY)
        mc(ws, R, 6, R, 7, value=f"=D{R}*E{R}", size=9, h="right", bdr=_b(), nf=MONEY)



    DE = R

    # Хэрэв employee байхгүй бол 1 хоосон мөр гаргая
    if DE < DS:
        R += 1
        DE = R
        ws.row_dimensions[R].height = 32
        for c in range(1, 7):
            ws.cell(R, c).border = _b()

    # ========= Totals =========
    R += 1
    ws.row_dimensions[R].height = 20
    mc(ws, R, 1, R, 5, value="Нийт төлбөр /НӨАТ ороогүй/", size=10, h="right", bdr=_b())
    mc(ws, R, 6, R, 7, value=f"=SUM(F{DS}:F{DE})", size=10, h="right", bdr=_b(), nf=MONEY)


    R += 1
    vat_row = R
    ws.row_dimensions[R].height = 20
    mc(ws, R, 1, R, 5, value="НӨАТ 10%", size=10, h="right", bdr=_b())
    mc(ws, R, 6, R, 7, value=f"=F{R-1}*0.10", size=10, h="right", bdr=_b(), nf=MONEY)

    R += 1
    ws.row_dimensions[R].height = 22
    mc(ws, R, 1, R, 5, value="Нийт төлбөр /НӨАТ орсон/", bold=True, size=10, h="right", bdr=_b())
    mc(ws, R, 6, R, 7, value=f"=F{vat_row-1}+F{vat_row}", bold=True, size=10, h="right", bdr=_b(), nf=MONEY)

    # Зураг дээрх шиг доор хар шугамтай хоосон том зай
    R += 1
    ws.row_dimensions[R].height = 42
    for c in range(1, 8):
        ws.cell(R, c).border = Border(top=_s("medium", BLACK), bottom=_s("medium", BLACK))

    # ========= Footer =========
    seller = company.get("seller_name", "Л. Анужин")
    
    R += 3

    # -------- Row 1 --------
    ws.row_dimensions[R].height = 24

    mc(
        ws, R, 1, R, 2,
        value="Борлуулагч:",
        size=9, wrap=True, h="left", v="center"
    )

    sc(ws, R, 3, "",h="left", bdr=_bm())

    mc(
        ws, R, 3, R, 4,
        value=seller,
        size=9, wrap=True, h="center", v="center"
    )

    sc(
        ws, R, 5,
        "Барааг хүргэсэн түгээгч:",
        size=9, wrap=True, h="left", v="center"
    )

    sc(ws, R, 6, "",h="left", bdr=_bm())
    sc(ws, R, 7, "",h="left", bdr=_bm())

    # -------- Row 2 --------
    R += 1
    ws.row_dimensions[R].height = 36

    mc(
        ws, R, 1, R, 2,
        value="Барааг олгосон нярав:",
        size=9, wrap=True, h="left", v="center"
    )

    sc(ws, R, 3, "",h="left", bdr=_bm())

    mc(
        ws, R, 3, R, 4,
        value="",
        size=9, h="center", v="center"
    )

    sc(
        ws, R, 5,
        "Бараа/ хайрцаг сав баглаа боодол\nбүрэн шалгаж/ хүлээн авсан:",
        size=9, wrap=True, h="left", v="center"
    )

    sc(ws, R, 6, "",h="left", bdr=_bm())
    sc(ws, R, 7, "",h="left", bdr=_bm())

    ws.print_area = f"A1:H{R}"


    return wb_to_pdf_bytes(wb)

INDIVIDUAL_EMPLOYEES = {"sergelen", "shajinbat"}  # lowercase-ээр харьцуулна

def strip_time_from_text(s: str) -> str:
    """Clock-style times болон илүүдэл зайг хасна (Excel/PDF тайлбарт)."""
    if not s:
        return s
    s = str(s)
    s = re.sub(r"\d{1,2}:\d{2}(?::\d{2})?(?:\s*[AP]M)?", "", s, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", s).strip()

def group_employees_for_invoice(employees: List[Dict]) -> List[Dict]:
    """
    - Sergelen / Shajinbat → тус бүр өөрийн invoice row
    - Бусад → level-ээр нэгтгэнэ (hours нийлнэ, unit_price ижил байна гэж үзнэ)
    """
    individual_rows = []
    level_groups: dict[str, dict] = {}  # key: level (e.g. "senior", "developer")

    for emp in employees:
        name_key = emp["employee_name"].strip().lower()
        if name_key in INDIVIDUAL_EMPLOYEES:
            # Тус тусдаа мөр — хэвээр нь авна
            individual_rows.append(emp)
        else:
            level = emp["level"].lower()  # "senior", "junior", "developer", etc.
            if level not in level_groups:
                level_groups[level] = {
                    "employee_name": f"K2 үйлчилгээний {level.title()}",
                    "level": level,
                    "total_hours": 0,
                    "unit_price": emp["unit_price"],  # ижил level = ижил үнэ гэж үзнэ
                    "cost_code": emp["cost_code"],
                }
            level_groups[level]["total_hours"] += emp["total_hours"]

    return individual_rows + list(level_groups.values())

def period_ym_from_doc_date(doc_date: str) -> Tuple[str, str]:
    """'2/25/2026' → ('2026', '02'). Алдаатай бол одоогийн сар."""
    doc_date = (doc_date or "").strip()
    y, m = "2026", "02"
    for sep in ("/", "-", "."):
        if sep in doc_date:
            parts = [p.strip() for p in doc_date.split(sep) if p.strip()]
            if len(parts) >= 3:
                if len(parts[0]) == 4:
                    y, m = parts[0], parts[1].zfill(2)
                else:
                    m, y = parts[0].zfill(2), parts[2]
                break
    return y, m

def k2_subheader_text(y: str, m: str, po_code: str) -> str:
    return f"Барааны нэр: K2 үйлчилгээний {y}/{m}-сар PO: {po_code}"

def k2_invoice_item_line(role_label: str, y: str, m: str) -> str:
    """Нэхэмжлэхийн барааны нэр — жишээ зурагтай ижил нэг мөр (K2)."""
    return strip_time_from_text(f"K2 үйлчилгээний {y} {m}-р сарын төлбөр {role_label}")

def split_excel_employees_for_k2(employees: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """Энгийн ажилтнууд (нийлбэр нэхэмжлэх) ба Shajinbat/Sergelen тусдаа PDF."""
    main, special = [], []
    for emp in employees:
        nk = emp["employee_name"].strip().lower()
        if nk in INDIVIDUAL_EMPLOYEES:
            special.append(emp)
        else:
            main.append(emp)
    return main, special

def aggregate_main_excel_by_senior_junior(
    employees: List[Dict], pricing: dict
) -> List[Dict]:
    """Senior эхлээд, дараа нь Developer/Junior — нэхэмжлэх/зарлагаийн мөрийн дараалал."""
    senior_h = 0.0
    junior_h = 0.0
    cc_senior = ""
    cc_junior = ""
    su = float(pricing.get("senior_unit_price", 32500))
    ju = float(pricing.get("junior_unit_price", 25200))
    for e in employees:
        lv = (e.get("level") or "junior").lower()
        h = float(e.get("total_hours") or 0)
        if lv == "senior":
            senior_h += h
            cc_senior = cc_senior or e.get("cost_code", "")
        else:
            junior_h += h
            cc_junior = cc_junior or e.get("cost_code", "")
    rows = []
    if senior_h > 0:
        rows.append({
            "employee_name": "Senior",
            "level": "senior",
            "total_hours": senior_h,
            "unit_price": su,
            "cost_code": cc_senior or "",
            "invoice_desc": None,
        })
    if junior_h > 0:
        rows.append({
            "employee_name": "Developer",
            "level": "junior",
            "total_hours": junior_h,
            "unit_price": ju,
            "cost_code": cc_junior or "",
            "invoice_desc": None,
        })
    return rows

def fill_k2_invoice_descriptions(rows: List[Dict], y: str, m: str, po_code: str) -> None:
    for e in rows:
        if e.get("invoice_desc"):
            e["invoice_desc"] = strip_time_from_text(e["invoice_desc"])
            continue
        if e.get("employee_name") in ("Senior", "Developer"):
            label = "Senior" if e.get("level") == "senior" else "Developer"
        elif e.get("employee_name"):
            nm = strip_time_from_text(e["employee_name"])
            label = f"{nm} ({'Senior' if e.get('level') == 'senior' else 'Developer'})"
        else:
            label = "Senior" if e.get("level") == "senior" else "Developer"
        e["invoice_desc"] = k2_invoice_item_line(label, y, m)

# ═══════════════════════════════════════════════════════════
# STREAMLIT UI
# ═══════════════════════════════════════════════════════════
st.set_page_config(page_title="Timesheet Processor", page_icon="🗓️", layout="wide")
st.title("Oyu Tolgoi Timesheet PDF generator v2ß")
st.markdown(
    "Upload **PDF** timesheets эсвэл **Excel** project list (.xlsx). "
    "Excel → **6 хуудасны PDF**: хуудас 1-2 бусад (Senior→Developer нэхэмжлэх+зарлага), "
    "хуудас 3-4 Shajinbat, хуудас 5-6 Sergelen."
)

# ── Load config ────────────────────────────────────────────
config_file = st.sidebar.file_uploader("⚙️ Upload employees_config.json", type=["json"])

if config_file:
    try:
        config = json.load(config_file)
        st.sidebar.success(f"✅ Config loaded: {len(config.get('employees', []))} employees")
    except Exception as e:
        st.sidebar.error(f"JSON parse error: {e}")
        st.stop()
else:
    try:
        config = load_config("employees_config.json")
        st.sidebar.info(f"📂 Using employees_config.json from disk ({len(config.get('employees', []))} employees)")
    except FileNotFoundError:
        st.sidebar.warning("⚠️ No config loaded — using PDF-parsed values & defaults")
        config = {"employees": [], "company": {}, "pricing": {}}

emp_lookup = build_employee_lookup(config)
name_lookup_by_name = build_employee_lookup_by_name(config)
pricing    = config.get("pricing", {})
company    = config.get("company", {})

if emp_lookup:
    with st.sidebar.expander(f"👥 {len(emp_lookup)} employees in config"):
        for sap, e in emp_lookup.items():
            st.write(f"**{e['employee_name']}** (SAP: {sap}) — {e.get('level','?')} — PO: {e.get('po_code','?')}")

# ══════════════════════════════════════════════════════════════════
# TAB LAYOUT
# ══════════════════════════════════════════════════════════════════
tab_ts, tab_proj = st.tabs(["📊 Data Team", "📊 Software Team"])

# ─────────────────────────────────────────────────────────────────
# TAB 1: Data Team — PDF timesheet → нэгтгэсэн PDF per PO
# ─────────────────────────────────────────────────────────────────
with tab_ts:
    uploaded_files = st.file_uploader(
        "📎 Upload PDF timesheet(s) эсвэл Excel (.xlsx)",
        type=["pdf", "xlsx"],
        accept_multiple_files=True,
    )

    if not uploaded_files:
        c1, c2 = st.columns(2)
        c1.info("**PDF:** олон ажилтан, хуудас бүр → нэхэмжлэх + зарлага")
        c2.info(
            "**Excel:** 6 хуудас — хуудас 1 бусад нэхэмжлэх (Senior→Developer), "
            "хуудас 2 зарлага, хуудас 3-4 Shajinbat, хуудас 5-6 Sergelen"
        )
    else:
        names = [f.name for f in uploaded_files]
        ext_set = {pathlib.Path(n).suffix.lower() for n in names}
        if len(ext_set) > 1:
            st.error("Нэг удаад зөвхөн PDF **эсвэл** зөвхөн Excel (.xlsx) оруулна уу.")
            st.stop()

        # ═══ Excel (К2 нэхэмжлэх) ═══
        if ".xlsx" in ext_set:
            po_groups_tsx = {}
            with st.spinner("Reading Excel…"):
                try:
                    for f in uploaded_files:
                        part = parse_project_list_excel(f.read())
                        for po, data in part.items():
                            if po not in po_groups_tsx:
                                po_groups_tsx[po] = {"employees": []}
                            po_groups_tsx[po]["employees"].extend(data["employees"])
                except Exception as e:
                    st.error(f"Excel уншихад алдаа: {e}")
                    st.stop()

            if not po_groups_tsx:
                st.warning("PO кодтой мөр олдсонгүй.")
            else:
                total_emps = sum(len(v["employees"]) for v in po_groups_tsx.values())
                total_hours = sum(
                    e["total_hours"] for v in po_groups_tsx.values() for e in v["employees"]
                )
                c1, c2, c3 = st.columns(3)
                c1.metric("PO codes", len(po_groups_tsx))
                c2.metric("Employee records", total_emps)
                c3.metric("Total hours", f"{total_hours:,.1f}")

                st.divider()
                st.subheader("⚙️ Баримтын тохиргоо (Excel / К2)")
                po_meta_tsx = {}
                for po, data in po_groups_tsx.items():
                    emp_names = ", ".join(e["employee_name"] for e in data["employees"])
                    with st.expander(
                        f"📋 PO **{po}** — {len(data['employees'])} record: {emp_names}",
                        expanded=True,
                    ):
                        ci1, ci2, ci3, ci4 = st.columns(4)
                        po_meta_tsx[po] = {
                            "inv_number": ci1.text_input(
                                "Нэхэмжлэх дугаар", key=f"tsx_inv_num_{po}", placeholder="26-OS02-010"
                            ),
                            "doc_date": ci2.text_input(
                                "Нэхэмжлэх огноо", key=f"tsx_inv_dt_{po}", placeholder="2/25/2026"
                            ),
                            "zar_number": ci3.text_input(
                                "Зарлагын баримт дугаар", key=f"tsx_zar_num_{po}", placeholder="26-OS01-008"
                            ),
                            "doc_date": ci4.text_input(
                                "Зарлагын огноо", key=f"tsx_zar_dt_{po}", placeholder="3/20/2026"
                            ),
                        }

                st.divider()
                all_tsx_pdfs: List[bytes] = []
                zip_tsx = io.BytesIO()
                with st.spinner("К2 PDF үүсгэж байна…"):
                    with zipfile.ZipFile(zip_tsx, "w") as zf:
                        for po, data in po_groups_tsx.items():
                            meta = po_meta_tsx[po]
                            for stem, blob in excel_po_k2_output_files(
                                po, data, company, pricing, name_lookup_by_name, meta
                            ):
                                all_tsx_pdfs.append(blob)
                                zf.writestr(f"{stem}.pdf", blob)
                    big_tsx = merge_pdfs(all_tsx_pdfs) if all_tsx_pdfs else b""

                st.subheader("⬇️ Татах")
                if big_tsx:
                    st.download_button(
                        label="📄 Бүх файлийг татах)",
                        data=big_tsx,
                        file_name="DataTeam_Excel_K2_Нэгтгэсэн.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="tsx_big_pdf",
                    )
                zip_tsx.seek(0)
                st.download_button(
                    "🗜️ Тусдаа файлууд (бусад / Shajinbat / Sergelen) ZIP",
                    data=zip_tsx.getvalue(),
                    file_name="DataTeam_K2_tusdaa_failuud.zip",
                    mime="application/zip",
                    key="tsx_zip",
                )
            st.stop()

        # ── Parse & enrich (PDF) ───────────────────────────────────
        all_emp = []
        with st.spinner("Reading PDFs…"):
            for f in uploaded_files:
                for emp in parse_pdf(f.read()):
                    emp["header"] = enrich_header(emp["header"], emp_lookup)
                    emp["header"]["unit_price"] = unit_price(emp["header"], pricing)
                    all_emp.append(emp)

        if not all_emp:
            st.error("No timesheet pages found.")
        else:
            emp_headers = [e["header"] for e in all_emp]
            po_groups   = group_by_po(emp_headers)
            unique_pos  = list(po_groups.keys())

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Files",           len(uploaded_files))
            m2.metric("Employees",       len(all_emp))
            m3.metric("Unique PO Codes", len(unique_pos))
            m4.metric("Total Hours",     sum(float(e["header"].get("total_hours") or 0) for e in all_emp))

            st.divider()
            st.subheader("⚙️ Invoice / Receipt settings per PO Code")

            po_meta = {}
            for po in unique_pos:
                members = po_groups[po]
                names   = ", ".join(h.get("employee_name","?") for h in members)
                with st.expander(f"📋 PO: **{po}** — {len(members)} employee(s): {names}", expanded=True):
                    table_rows = [
                        {
                            "Employee": e["employee_name"],
                            "Level": e["level"].title(),
                            "Hours": e["total_hours"],
                            "Unit Price ₮": f"{e['unit_price']:,.0f}",
                            "Total ₮": f"{e['total_hours'] * e['unit_price']:,.0f}",
                            "Cost Code": e["cost_code"],
                        }
                        for e in members
                    ]
                    st.dataframe(table_rows, use_container_width=True, hide_index=True)

                    ci1, ci2, ci3, ci4 = st.columns(4)
                    po_meta[po] = {
                        "inv_number": ci1.text_input(
                            "Нэхэмжлэх дугаар",
                            key=f"pl_inv_num_{po}",
                            placeholder="26-OS02-006",
                        ),
                        "doc_date": ci2.text_input(
                            "Нэхэмжлэх огноо",
                            key=f"pl_inv_dt_{po}",
                            placeholder="2/25/2026",
                        ),
                        "zar_number": ci3.text_input(
                            "Зарлагын баримт дугаар",
                            key=f"pl_zar_num_{po}",
                            placeholder="26-OS01-008",
                        ),
                        "doc_date": ci4.text_input(
                            "Зарлагын баримт огноо",
                            key=f"pl_zar_dt_{po}",
                            placeholder="3/20/2026",
                        ),
                    }

            st.divider()

            # ── Бүх ажилтны PDF-ийг нэг том PDF болгоно ──
            # Дараалал: Ажилтан 1 (Нэхэмжлэх + Зарлагын баримт + Timesheet)
            #           Ажилтан 2 (Нэхэмжлэх + Зарлагын баримт + Timesheet) ...
            all_pdfs = []
            with st.spinner("Generating & merging all PDFs into one…"):
                for emp in all_emp:
                    h    = emp["header"]
                    po   = h.get("po_code", "UNKNOWN")
                    meta = po_meta.get(po, {
                        "inv_number": "", "doc_date": "",
                        "zar_number": "", "doc_date": "",
                    })

                    inv_pdf = build_invoice(
                        [h], pricing, company,
                        meta["inv_number"], meta["doc_date"])

                    zar_pdf = build_zarlagiin(
                        [h], pricing, company,
                        meta["zar_number"], meta["doc_date"])

                    # Нэг ажилтны 2 документ дараалан нэмнэ (нэхэмжлэх + зарлага)
                    all_pdfs.extend([inv_pdf, zar_pdf])

            big_pdf = merge_pdfs(all_pdfs)

            # ── Download ─────────────────────────────────────────────
            st.subheader("⬇️ Download Documents")
            st.download_button(
                label="📄 Бүх ажилтны нэгтгэсэн PDF татах",
                data=big_pdf,
                file_name="Бүх_ажилтан_Нэгтгэсэн.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="big_pdf_download",
            )


            # ── ZIP: нэг том PDF + xlsx ───────────────────────────────
            st.divider()
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w") as zf:
                zf.writestr("Бүх_ажилтан_Нэгтгэсэн.pdf", big_pdf)
                for emp in all_emp:
                    h      = emp["header"]
                    name   = h.get("employee_name") or "Employee"
                    period = (h.get("period_start") or "").replace(".", "-")
                    prefix = safe(name)
            zip_buf.seek(0)
            st.download_button(
                "🗜️ Download All (ZIP)",
                data=zip_buf,
                file_name="Timesheets_All.zip",
                mime="application/zip",
            )

            # ── Employee detail tabs ─────────────────────────────────
            st.divider()
            st.subheader("👥 Employee Details")
            tabs = st.tabs([e["header"].get("employee_name") or f"Emp {i+1}"
                            for i, e in enumerate(all_emp)])
            for tab, emp in zip(tabs, all_emp):
                h, rows = emp["header"], emp["rows"]
                with tab:
                    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
                    c1.metric("SAP",         h.get("sap_number") or "—")
                    c2.metric("Level",       h.get("level", "—").title())
                    c3.metric("Work Days",   h.get("work_days") or "—")
                    c4.metric("Total Hours", h.get("total_hours") or "—")
                    c5.metric("Cost Code",   h.get("cost_code") or "—")
                    c6.metric("PO Code",     h.get("po_code") or "—")
                    c7.metric("Daily Rows",  len(rows))
                    st.caption(
                        f"**Period:** {h.get('period_start','')} → {h.get('period_end','')}   |   "
                        f"**Position:** {h.get('position') or '—'}   |   "
                        f"**OT Leader:** {h.get('ot_leader_name') or '—'}   |   "
                        f"**Product Code:** {h.get('product_code') or '—'}"
                    )


# ─────────────────────────────────────────────────────────────────
# TAB 2: Software Team — Excel project list → нэгтгэсэн PDF per PO
# ─────────────────────────────────────────────────────────────────
with tab_proj:
    st.markdown(
        "Upload the **ConT/VWL Project List** Excel. "
        "Generates a **6-page PDF per PO**: "
        "page 1 main invoice (Senior→Developer), page 2 main zarlaga, "
        "page 3-4 Shajinbat invoice+zarlaga, page 5-6 Sergelen invoice+zarlaga. "
        "Clock times are stripped from all descriptions."
    )

    proj_file = st.file_uploader(
        "📎 Upload Project List Excel (.xlsx)",
        type=["xlsx"],
        key="proj_list_upload",
    )

    if not proj_file:
        st.info("No file uploaded yet. Upload the `_ConT___VWL___Report_*.xlsx` file above.")
    else:
        with st.spinner("Reading Excel…"):
            try:
                proj_file_bytes = proj_file.read()
                po_groups_proj = parse_project_list_excel(proj_file_bytes)
            except Exception as e:
                st.error(f"Failed to parse Excel: {e}")
                st.stop()

        if not po_groups_proj:
            st.warning("No valid PO code rows found in the uploaded file.")
        else:
            total_emps = sum(len(v["employees"]) for v in po_groups_proj.values())
            total_hours = sum(
                emp["total_hours"]
                for v in po_groups_proj.values()
                for emp in v["employees"]
            )

            c1, c2, c3 = st.columns(3)
            c1.metric("Unique PO Codes", len(po_groups_proj))
            c2.metric("Employee Records", total_emps)
            c3.metric("Total Hours", f"{total_hours:,.1f}")

            st.divider()
            st.subheader("⚙️ Document settings per PO Code")

            po_meta_proj = {}
            from collections import defaultdict

            # 1. PO-г 3 group болгох
            grouped_po = {
                "3104799438": {"employees": []},
                "3104801353": {"employees": []},
                "other": {"employees": []},
            }

            for po, data in po_groups_proj.items():
                target_group = po if po in ["3104799438", "3104801353"] else "other"
                grouped_po[target_group]["employees"].extend(data["employees"])


            # 2. UI render
            po_meta_proj = {}

            for po_group, data in grouped_po.items():
                if not data["employees"]:
                    continue

                emp_names = ", ".join(e["employee_name"] for e in data["employees"])

                with st.expander(
                    f"📋 PO: **{po_group}** — {len(data['employees'])} employee(s): {emp_names}",
                    expanded=True,
                ):
                    table_rows = [
                        {
                            "Employee": e["employee_name"],
                            "Level": e["level"].title(),
                            "Hours": e["total_hours"],
                            "Unit Price ₮": f"{e['unit_price']:,.0f}",
                            "Total ₮": f"{e['total_hours'] * e['unit_price']:,.0f}",
                            "Cost Code": e["cost_code"],
                        }
                        for e in data["employees"]
                    ]

                    st.dataframe(table_rows, use_container_width=True, hide_index=True)

                    # 3. 3 field + 1 merged date field
                    ci1, ci2, ci3 = st.columns(3)

                    po_meta_proj[po_group] = {
                        "inv_number": ci1.text_input(
                            "Нэхэмжлэх дугаар",
                            key=f"pl_inv_num_{po_group}",
                            placeholder="26-OS02-006",
                        ),
                        "zar_number": ci2.text_input(
                            "Зарлагын баримт дугаар",
                            key=f"pl_zar_num_{po_group}",
                            placeholder="26-OS01-008",
                        ),
                        "doc_date": ci3.text_input(
                            "Огноо",
                            key=f"pl_doc_dt_{po_group}",
                            placeholder="2/25/2026",
                        ),
                    }
            
            st.divider()
            st.subheader("⬇️ Үүсгэх & татах")

            # Reset generated files when a new file is uploaded
            current_proj_file_name = proj_file.name if proj_file else None
            if "proj_last_uploaded_file" not in st.session_state:
                st.session_state.proj_last_uploaded_file = None
            if "proj_generated_files" not in st.session_state:
                st.session_state.proj_generated_files = None

            if st.session_state.proj_last_uploaded_file != current_proj_file_name:
                st.session_state.proj_last_uploaded_file = current_proj_file_name
                st.session_state.proj_generated_files = None

            generate_proj_pdf = st.button(
                "📄 PDFs үүсгэх",
                use_container_width=True,
                key="generate_proj_pdf_btn",
            )

            if generate_proj_pdf:
                all_proj_pdfs: List[bytes] = []
                zip_proj = io.BytesIO()

                with st.spinner("К2 багийн PDF үүсгэж байна…"):
                    with zipfile.ZipFile(zip_proj, "w") as zf:
                        for po, data in po_groups_proj.items():
                            po_group = po if po in ["3104799438", "3104801353"] else "other"
                            meta = po_meta_proj.get(po_group)

                            if not meta:
                                st.warning(f"Metadata not found for PO {po} (group: {po_group})")
                                continue

                            for stem, blob in excel_po_k2_output_files(
                                po, data, company, pricing, name_lookup_by_name, meta
                            ):
                                all_proj_pdfs.append(blob)
                                zf.writestr(f"{stem}.pdf", blob)

                    big_proj_pdf = merge_pdfs(all_proj_pdfs) if all_proj_pdfs else b""
                    zip_proj.seek(0)

                    st.session_state.proj_generated_files = {
                        "big_pdf": big_proj_pdf,
                        "zip_bytes": zip_proj.getvalue(),
                    }

                st.success("PDF үүсгэлээ! Доор татаж авна уу. Ануууууууу")

            if st.session_state.proj_generated_files:
                st.download_button(
                    label="📄 Бүх файлыг татах",
                    data=st.session_state.proj_generated_files["big_pdf"],
                    file_name="Бүх_ажилтан_ProjectList_K2_Нэгтгэсэн.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="big_proj_pdf_download",
                )

                st.divider()

                st.download_button(
                    "🗜️ Тусдаа файлууд ZIP",
                    data=st.session_state.proj_generated_files["zip_bytes"],
                    file_name="ProjectList_K2_tusdaa_failuud.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key="proj_zip_download",
                )